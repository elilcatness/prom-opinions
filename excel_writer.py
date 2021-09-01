from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


class ExcelWriter:
    filename: str
    sheet_names: list
    headers: dict

    def __init__(self, filename, sheet_names=None, headers=None):
        self.workbook = Workbook()
        self.filename = filename
        if sheet_names is None:
            sheet_names = self.workbook.sheetnames
        if not isinstance(sheet_names, list):
            if sheet_names is None:
                sheet_names = []
            else:
                raise TypeError('sheet_names must be a list')
        elif sheet_names != self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        for i, sheet in enumerate(sheet_names):
            if sheet not in self.workbook.sheetnames:
                self.workbook.create_sheet(sheet, i)
        if not isinstance(headers, dict):
            if isinstance(headers, list):
                headers = {'Sheet': headers}
            elif headers is None:
                headers = {'Sheet': []}
            else:
                raise TypeError('headers must be dict')
        self.check_sheets(headers.keys(), remark='headers keys')
        self.headers = headers

    def check_sheets(self, sheet_list, remark=None):
        """
        :param sheet_list: Sheet names list for checking whether ones are in the file's sheets
        :param remark: Origin of given sheet list
        """
        unknown_sheets = list(filter(lambda sh: sh not in self.workbook.sheetnames, sheet_list))
        if unknown_sheets:
            raise NameError(f'{", ".join(unknown_sheets)} not in sheets of {self.filename}'
                            if remark is None else
                            f'{", ".join(unknown_sheets)} ({remark}) not in sheets of {self.filename}')

    def check_headers(self, sheet_name, headers, remark=None):
        """
        :param sheet_name: Name of sheet in which headers the given one is checked for being in
        :param headers: List of headers for checking whether ones are in given sheet's headers
        :param remark: Origin of given headers list
        """
        self.check_sheets([sheet_name])
        unknown_headers = list(filter(lambda header: header not in self.headers[sheet_name],
                                      headers))
        if unknown_headers:
            raise NameError(f'{", ".join(unknown_headers)} not in headers of sheet {sheet_name}'
                            if remark is None else
                            f'{", ".join(unknown_headers)} ({remark}) '
                            f'not in sheets of sheet {sheet_name}')

    def write_headers(self, sheet_names=None, bold=False, auto_save=True):
        if not sheet_names:
            sheet_names = [self.workbook.active.title]
        if not isinstance(sheet_names, list):
            sheet_names = [sheet_names]
        self.check_sheets(sheet_names)
        for sheet_name in sheet_names:
            sheet: Worksheet = self.workbook[sheet_name]
            headers = self.headers[sheet_name]
            for i in range(0, len(headers)):
                cell = sheet.cell(1, i + 1, headers[i])
                if bold:
                    cell.font = Font(bold=True)
        if auto_save:
            self.workbook.save(self.filename)

    def write_row(self, data, sheet_name=None, auto_save=True):
        if not sheet_name:
            sheet_name = self.workbook.active.title
        self.check_headers(sheet_name, data.keys())
        sheet: Worksheet = self.workbook[sheet_name]
        row = sheet.max_row + 1
        for i, items in enumerate(data.items()):
            key, val = items
            sheet.cell(row, self.headers[sheet_name].index(key) + 1, val)
        if auto_save:
            self.workbook.save(self.filename)
